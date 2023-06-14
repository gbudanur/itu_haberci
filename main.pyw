from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.chrome.options import Options
from openpyxl import Workbook, load_workbook
import os
from time import sleep
import requests
from win10toast import ToastNotifier
import json

directory = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(directory, "database.xlsx")
login_url = "https://ninova.itu.edu.tr/Kampus1"
options = Options()
options.add_argument("--headless")
t_tech = 3
t = 900


def show_notification(title, message):
    toaster = ToastNotifier()
    toaster.show_toast(title, message, duration=0)


def email(konu, coursename, gradename):
    return requests.post(
        "https://api.mailgun.net/v3/" + ws["A4"].value + "/messages",
        auth=("api", ws["B4"].value),
        data={
            "from": "Haberci <ituhaberci@" + ws["A4"].value + ">",
            "to": [ws["A6"].value, "Kullanıcı"],
            "subject": konu,
            "template": "notification",
            "t:variables": json.dumps({"CourseName": coursename, "Grade": gradename}),
        },
    )


def notify(coursename, gradename, grade):
    konu = coursename + "  dersinize yeni not eklendi. " + gradename + " " + grade
    text = (
        "Merhaba, "
        + coursename
        + " dersinize yeni not eklendi. "
        + gradename
        + " "
        + grade
        + " \n\n ituhaberci.com"
    )
    coursename = coursename
    gradename = gradename + " " + grade
    email(konu, coursename, gradename)
    show_notification(konu, text)
    print("Bildiri e-mail'i atıldı.")


def init():
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    driver.get(login_url)
    sleep(t_tech)
    driver.find_element(
        By.XPATH, "/html/body/form/div[5]/div/div/div[1]/div/ul/li/div[1]/input"
    ).send_keys(ws["A2"].value)
    driver.find_element(
        By.XPATH, "/html/body/form/div[5]/div/div/div[1]/div/ul/li/div[2]/input"
    ).send_keys(ws["B2"].value)
    driver.find_element(
        By.XPATH, "/html/body/form/div[5]/div/div/div[1]/div/ul/li/input"
    ).click()
    sleep(t_tech)
    i = 4
    while ws.cell(row=1, column=i).value is not None:
        driver.get(ws.cell(row=1, column=i).value)
        sleep(t_tech)

        for j in range(2, 100):
            try:
                grade = driver.find_element(
                    By.XPATH,
                    "/html/body/form/div[3]/div[3]/div[3]/div/table/tbody/tr["
                    + str(j)
                    + "]/td[2]/span",
                ).text
                ws.cell(row=j, column=i, value=grade)
            except NoSuchElementException:
                break
        i += 1
    wb.save(file_path)
    driver.quit()


def update():
    driver = webdriver.Chrome(ChromeDriverManager().install(), options=options)
    wb = load_workbook(file_path)
    ws = wb.active
    driver.get(login_url)
    sleep(t_tech)
    driver.find_element(
        By.XPATH, "/html/body/form/div[5]/div/div/div[1]/div/ul/li/div[1]/input"
    ).send_keys(ws["A2"].value)
    driver.find_element(
        By.XPATH, "/html/body/form/div[5]/div/div/div[1]/div/ul/li/div[2]/input"
    ).send_keys(ws["B2"].value)
    driver.find_element(
        By.XPATH, "/html/body/form/div[5]/div/div/div[1]/div/ul/li/input"
    ).click()
    sleep(t_tech)
    i = 4

    while ws.cell(row=1, column=i).value is not None:
        driver.get(ws.cell(row=1, column=i).value)
        sleep(t_tech)
        for j in range(2, 100):
            try:
                grade = driver.find_element(
                    By.XPATH,
                    "/html/body/form/div[3]/div[3]/div[3]/div/table/tbody/tr["
                    + str(j)
                    + "]/td[2]/span",
                ).text
                if grade != ws.cell(row=j, column=i).value and grade != "":
                    coursename = driver.find_element(
                        By.XPATH, "/html/body/form/div[3]/div[2]/div[1]/a[3]"
                    ).text
                    gradename = driver.find_element(
                        By.XPATH,
                        "/html/body/form/div[3]/div[3]/div[3]/div/table/tbody/tr["
                        + str(j)
                        + "]/td[1]/span",
                    ).text
                    ws.cell(row=j, column=i, value=grade)
                    notify(coursename, gradename, grade)

            except NoSuchElementException:
                break
        i += 1
    wb.save(file_path)
    driver.quit()


try:
    wb = load_workbook(file_path)
    ws = wb.active
    newstartup = False

except FileNotFoundError:
    newstartup = True
    wb = Workbook()
    ws = wb.active
    username = input("Kullanıcı adınızı girin: ")
    password = input("Şifrenizi girin: ")
    api_key = input("Mailgun hesabınıza tanımlanan API_KEY'i girin: ")
    api_url = input("Mailgun hesabınıza tanımlanan API_DOMAIN'i girin: ")
    receiver = input("Bildirimlerin gönderileceği mail adresini girin: ")
    ws["A1"] = "username"
    ws["B1"] = "password"
    ws["A2"] = username
    ws["B2"] = password
    ws["A3"] = "api_url"
    ws["B3"] = "api_key"
    ws["A4"] = api_url
    ws["B4"] = api_key
    ws["A5"] = "receiver"
    ws["A6"] = receiver

    print(
        "Kontrol edilmesini istediğiniz notların yayınlandığı url'leri tek tek girin, çıkmak için 0 yazın. Örnek url: https://ninova.itu.edu.tr/Sinif/11111.11111/Notlar"
    )
    k = 4

    while True:
        url = input(": ")
        if url == "0":
            break
        ws.cell(row=1, column=k, value=url)
        k += 1

    init()

while True:
    if newstartup:
        newstartup = False
        print("Sleeping for 30 minutes...")
        sleep(t)
    update()
    print("Sleeping for 30 minutes...")
    sleep(t)
